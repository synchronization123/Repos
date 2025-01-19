import os
import requests
import pandas as pd
from io import StringIO
import tkinter as tk
from tkinter import ttk, simpledialog, messagebox
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
import json

# URL to fetch tokens from
TOKEN_URL = "https://demo.defectdojo.org/media/uploaded_files/348c0cb6-11a8-4735-8d3d-61cbd2535d45.json"



# Function to load tokens from the remote JSON file
def load_tokens():
    try:
        response = requests.get(TOKEN_URL)
        response.raise_for_status()  # Check if the request was successful
        tokens = response.json()
        return tokens.get("jira_token", ""), tokens.get("dojo_token", "")
    except requests.RequestException as e:
        console.insert(tk.END, f"❌ Error fetching tokens: {e}\n")
        return '', ''  # Return empty tokens in case of error

# Function to save tokens to the remote JSON file (optional)
def save_tokens(jira_token, dojo_token):
    # Normally you'd want to save this remotely too, but for this example, we aren't updating the remote file.
    console.insert(tk.END, "⚠️ Tokens are saved locally for this session.\n")

# Initial tokens (loaded from the remote URL)
jira_token, dojo_token = load_tokens()

# Function to download usernames
def download_usernames(progress_var):
    console.delete("1.0", tk.END)  # Clear previous logs
    progress_var.set(10)
    root.update_idletasks()
    
    url = "https://demo.defectdojo.org/api/v2/users/?offset=0&limit=100"
    headers = {'Authorization': f'Token {dojo_token}'}
    
    try:
        response = requests.get(url, headers=headers)
        response.raise_for_status()
        data = response.json()
        
        usernames = [(user['id'], user['username']) for user in data['results']]
        df = pd.DataFrame(usernames, columns=['ID', 'Username'])
        df.to_excel("usernames.xlsx", index=False)
        
        console.insert(tk.END, "✅ Script started loading prerequisites.\n")
    
    except requests.RequestException as e:
        console.insert(tk.END, f"❌ Error loading prerequisites: {e}\n")
    
    progress_var.set(30)
    root.update_idletasks()

# Function to fetch JIRA data with pagination
def fetch_jira_data(progress_var, jql_query):
    url = f"https://jira.demo.almworks.com/sr/jira.issueviews:searchrequest-csv-all-fields/temp/SearchRequest.csv?jqlQuery={jql_query}"
    headers = {'Authorization': f'Token {jira_token}'}
    offset = 0
    limit = 1000  # Maximum JIRA allows per request
    all_data = []  # To store all the data from paginated requests
    
    try:
        while True:
            response = requests.get(f"{url}&offset={offset}&limit={limit}", headers=headers)
            
            if response.status_code == 400:
                # If the request fails with status 400, delete the usernames file
                if os.path.exists("usernames.xlsx"):
                    os.remove("usernames.xlsx")
                    console.insert(tk.END, "❌ JIRA API request failed. Status: 400 \n")
                break
            
            if response.status_code != 200:
                console.insert(tk.END, f"❌ JIRA API request failed. Status: {response.status_code}\n")
                break
            
            csv_data = response.text
            if not csv_data.strip():  # If no data is returned
                console.insert(tk.END, "❌ No data found for the JQL query.\n")
                break
            
            # Parse the CSV data
            data = pd.read_csv(StringIO(csv_data))
            all_data.append(data)  # Append current batch of data
            
            # If the batch has fewer rows than the limit, we've reached the end
            if len(data) < limit:
                break

            # Move to the next batch (next 1000 results)
            offset += limit
        
        # Concatenate all data into one DataFrame
        merged_data = pd.concat(all_data, ignore_index=True)

        # Extract relevant columns
        comment_cols = [col for col in merged_data.columns if "Comment" in col]
        label_cols = [col for col in merged_data.columns if "Labels" in col]

        # Merge all label columns into one, removing duplicates
        merged_data['Labels'] = merged_data[label_cols].apply(lambda row: ', '.join(sorted(set(row.dropna().astype(str)))), axis=1)

        # Add Category column based on Labels content
        merged_data['Category'] = merged_data['Labels'].apply(lambda x: 'Security Jira' if 'Security' in x or 'AppSecurity' in x else 'Functional Jira')

        # Keep only relevant columns
        merged_data = merged_data[["Issue key", "Summary", "Labels", "Category"] + comment_cols]

        # Save to files
        merged_data.to_csv("jiradata.csv", index=False)
        merged_data.to_excel("jiradata.xlsx", index=False)
        
        console.insert(tk.END, "✅ JIRA data fetched successfully.\n")
    
    except Exception as e:
        console.insert(tk.END, f"❌ Error processing JIRA data: {e}\n")
    
    progress_var.set(60)
    root.update_idletasks()

# Function to highlight usernames, reformat comments, and style Issue Key in red
def highlight_usernames_and_reformat_comments(progress_var):
    try:
        usernames_df = pd.read_excel("usernames.xlsx")
        usernames = set(usernames_df['Username'].astype(str).str.lower())
        
        wb = load_workbook("jiradata.xlsx")
        ws = wb.active
        
        yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        orange_font = Font(color="FFA500")  # Orange color for usernames
        red_font = Font(color="FF0000")    # Red color for Issue Key
        wrap_text = Alignment(wrap_text=True)
        
        comment_cols = [cell.column_letter for cell in ws[1] if "Comment" in cell.value]

        for row in ws.iter_rows(min_row=2):
            issue_key_found = False  # To check if any username is found in comments
            
            for cell in row:
                if cell.column_letter in comment_cols and cell.value:
                    # Split the comment and reformat it
                    comment_parts = cell.value.split(';')
                    if len(comment_parts) >= 3:
                        date_time = comment_parts[0].strip()
                        username = comment_parts[1].strip()
                        comment_text = comment_parts[2].strip()

                        # Limit comment to the first 4 lines (if more than 4 lines)
                        comment_lines = comment_text.splitlines()[:4]
                        limited_comment = '\n'.join(comment_lines)

                        # Reformat comment
                        formatted_comment = (
                            f"Date and Time: {date_time}\n"
                            f"Comment Added by: {username}\n"
                            f"Comment: {limited_comment}"
                        )

                        # Set the formatted comment
                        cell.value = formatted_comment
                        cell.alignment = wrap_text

                        # Highlight username in orange
                        if username.lower() in usernames:
                            issue_key_found = True  # If the username is found in the comments
                            cell.font = Font(color="FFA500")  # Orange color

            # Highlight Issue Key in red if no username is found in comments
            issue_key_cell = row[0]  # Assuming Issue key is in the first column
            if not issue_key_found:
                issue_key_cell.font = red_font  # Make Issue Key red
            
        # Set fixed row height to 141.75 points (5 cm)
        for row in ws.iter_rows():
            ws.row_dimensions[row[0].row].height = 40.75  # 5 cm = 141.75 points

        # Adjust column width
        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = 40  

        wb.save("Report.xlsx")
        console.insert(tk.END, "✅ Report generated successfully.\n")
        
        # After generating the report, delete unnecessary files
        delete_files()
    
    except Exception as e:
        console.insert(tk.END, f"❌ Error generating report \n")
    
    progress_var.set(100)
    root.update_idletasks()

# Function to delete unnecessary files
def delete_files():
    files_to_delete = ["usernames.xlsx", "jiradata.csv", "jiradata.xlsx"]
    
    for file in files_to_delete:
        if os.path.exists(file):
            os.remove(file)
            
# Function to trigger the whole process
def process_data():
    progress_var.set(0)
    root.update_idletasks()

    # Get the JQL entered by the user
    jql_query = entry_jql.get().strip()

    # Handle optional JQL2 (keyword in comments)
    keyword = entry_keyword.get().strip()
    if keyword:
        jql_query += f' AND comment~"{keyword}"'

    # Handle optional JQL3 (labels)
    labels = entry_labels.get().strip()
    if labels:
        jql_query += f' AND labels={labels}'

    # If no JQL query is entered, notify the user
    if not jql_query:
        console.insert(tk.END, "❌ No JQL conditions entered.\n")
        return
    
    # Trigger the steps to download usernames, fetch JIRA data, and highlight comments
    download_usernames(progress_var)
    fetch_jira_data(progress_var, jql_query)
    highlight_usernames_and_reformat_comments(progress_var)
    delete_files()  # Delete unnecessary files at the end


# Main application setup
root = tk.Tk()
root.title("JIRA Report Generator")

frame = ttk.Frame(root, padding="10")
frame.grid(row=0, column=0)

ttk.Label(frame, text="Enter JQL Query (Mandatory):", anchor='w').grid(row=0, column=0, sticky='w', pady=5)
entry_jql = ttk.Entry(frame, width=50)
entry_jql.grid(row=0, column=1, pady=5)

ttk.Label(frame, text="Search Keyword in Comment (Optional):", anchor='w').grid(row=1, column=0, sticky='w', pady=5)
entry_keyword = ttk.Entry(frame, width=50)
entry_keyword.grid(row=1, column=1, pady=5)

ttk.Label(frame, text="Search by Label (Optional):", anchor='w').grid(row=2, column=0, sticky='w', pady=5)
entry_labels = ttk.Entry(frame, width=50)
entry_labels.grid(row=2, column=1, pady=5)

# Progress bar
progress_var = tk.DoubleVar()
progress_bar = ttk.Progressbar(frame, variable=progress_var, maximum=100)
progress_bar.grid(row=3, column=0, columnspan=2, pady=10)

# Process button
process_button = ttk.Button(frame, text="Generate Report", command=process_data)
process_button.grid(row=4, column=0, columnspan=2, pady=10)

# Console output area
console = tk.Text(frame, width=80, height=20)
console.grid(row=5, column=0, columnspan=2, pady=10)

root.mainloop()
